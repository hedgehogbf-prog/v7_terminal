# mppt/logger.py — объединённая версия:
# - старый рабочий парсинг (bracket/number + правильные цвета)
# - новая логика auto-PASSED, Git-интеграция и статус-бары
# - фикс TXT-лога и автоконфигурации Git

from __future__ import annotations

import os
import re
import subprocess
import time
from typing import Callable, Optional, Tuple, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from util.fileutil import get_log_paths, timestamp_str
from util.ansi import strip_ansi
from mppt.terminal_pyte import PYTE_FG_TO_HEX


def _excel_color_from_hex(term_hex: str) -> str:
    """
    Преобразует цвет из CanvasTerminal ("#RRGGBB") в цвет Excel ("RRGGBB")
    по схеме:
    - зелёный  -> 00AA00
    - красный  -> FF0000
    - всё остальное -> 000000 (чёрный)
    """
    if not term_hex:
        return "000000"
    hex_clean = term_hex.lstrip("#").lower()

    green_set = {
        PYTE_FG_TO_HEX["green"].lstrip("#").lower(),
        PYTE_FG_TO_HEX.get("brightgreen", "").lstrip("#").lower(),
    }
    red_set = {
        PYTE_FG_TO_HEX["red"].lstrip("#").lower(),
        PYTE_FG_TO_HEX.get("brightred", "").lstrip("#").lower(),
    }

    if hex_clean in green_set:
        return "00AA00"
    if hex_clean in red_set:
        return "FF0000"
    return "000000"


class MPPTLogger:
    """
    Логгер для MPPT:

    - ведёт txt-лог кадра (сырой текст без ANSI)
    - пишет в Excel:
        * основной лист "Sheet" — все сохранения
        * лист "PASSED"        — кадры, где есть PASSED
    - защита от повторов: в одном сеансе для одного и того же ID автозапись
      PASSED делается только один раз
    - Git-интеграция:
        * git pull при старте (если каталог логов — git-репозиторий)
        * git add/commit по кнопке
        * git push по кнопке
      Статусы Git выводятся через git_status_callback, остальное — через status_callback.
    """

    def __init__(
        self,
        base_dir: Optional[str] = None,
        status_callback: Optional[Callable[[str, str], None]] = None,
    ):
        # пути к txt и xlsx логам
        self.txt_path, self.xlsx_path = get_log_paths(base_dir)

        # общий статус (главный status bar приложения)
        self.status_callback: Optional[Callable[[str, str], None]] = status_callback
        # отдельный Git-статус (заполняется в MPPTTerminalPanel)
        self.git_status_callback: Optional[Callable[[str, str], None]] = None

        # запоминаем, для какого ID уже автозаписывали PASSED в этом сеансе
        self.last_passed_id: Optional[str] = None

        # каталог, где лежат логи — там же ожидаем git-репозиторий
        self.logs_dir = os.path.dirname(self.txt_path) or os.getcwd()

    # ----------------------------------------------------------
    # Статусы
    # ----------------------------------------------------------
    def _set_status(self, msg: str, color: str = "white") -> None:
        """Общий статус (COM, Excel и т.п.)."""
        if self.status_callback:
            self.status_callback(msg, color)
        else:
            print(msg)

    def _set_git_status(self, msg: str, color: str = "white") -> None:
        """Отдельный статус для Git. Если git_status_callback не задан, используем общий статус."""
        if self.git_status_callback:
            try:
                self.git_status_callback(msg, color)
                return
            except Exception:
                # если что-то пошло не так с callback — не валимся, а пишем в общий статус/консоль
                pass
        self._set_status(msg, color)

    # ----------------------------------------------------------
    # Excel
    # ----------------------------------------------------------
    def _ensure_workbook(self) -> Tuple[Workbook, object, object]:
        """
        Создаёт/открывает Excel-файл.
        Возвращает три листа:
            wb, основной ws, PASSED-лист ws_passed
        """
        # если файл есть — пробуем открыть
        if os.path.exists(self.xlsx_path):
            try:
                wb = load_workbook(self.xlsx_path)

                # основной лист
                ws = wb.active
                if ws.title != "Sheet" and "Sheet" in wb.sheetnames:
                    ws = wb["Sheet"]

                # PASSED-лист
                if "PASSED" in wb.sheetnames:
                    ws_passed = wb["PASSED"]
                else:
                    ws_passed = wb.create_sheet("PASSED")
                    # создаём шапку, если в основном листе она есть
                    if ws.max_row >= 1:
                        header = [cell.value for cell in ws[1]]
                        ws_passed.append(header)

                return wb, ws, ws_passed

            except Exception as e:
                # файл битый — переименуем и создадим новый
                self._set_status(
                    f"MPPT: повреждён Excel-лог, создаю новый: {e}", "yellow"
                )
                try:
                    backup = self.xlsx_path + ".corrupt"
                    os.replace(self.xlsx_path, backup)
                except OSError:
                    pass

        # если файла нет или он был битый — создаём новый
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"

        header = [
            "ID",
            "UART",
            "Voltage",
            "U_bat",
            "U_src",
            "Current",
            "I_crg",
            "I_ch1",
            "I_ch2",
            "Charger",
            "M_sens",
            "L_sens",
        ]
        ws.append(header)

        ws_passed = wb.create_sheet("PASSED")
        ws_passed.append(header)

        os.makedirs(os.path.dirname(self.xlsx_path), exist_ok=True)
        wb.save(self.xlsx_path)
        return wb, ws, ws_passed

    # ----------------------------------------------------------
    # Цвет строки
    # ----------------------------------------------------------
    def _row_hex_color(self, color_matrix, row_idx: int) -> Optional[str]:
        """
        Берёт цвет первой видимой цветной ячейки строки pyte.
        Устойчив к любой длине матрицы и пустым строкам.
        """
        if color_matrix is None:
            return None
        if not (0 <= row_idx < len(color_matrix)):
            return None

        row = color_matrix[row_idx]
        if not row:
            return None

        for c in row:
            if c:
                return c
        return None

    # ----------------------------------------------------------
    # Парсинг кадра
    # ----------------------------------------------------------
    def _parse_frame(
        self, lines: List[str], color_matrix
    ) -> Tuple[List[str], List[str]]:
        """
        Парсинг кадра (произвольное число строк pyte.get_lines()) в массив значений и цветов.
        Формат строк (пример):

            STM32f030x     ID:BCFB
            UART            [++++++]
            Voltage         [+]
             U_bat  14002   mV
             U_src  15011   mV
            Current         [+]
             I_crg  4       mA
             I_ch1  3       mA
             I_ch2  4       mA
            Charger         [+]
            M_sens          [+]
            L_sens          [+]
            [-PASSED-]
        """
        values = ["" for _ in range(12)]
        colors = ["000000" for _ in range(12)]

        plain_lines = [strip_ansi(l) for l in lines]

        # ---------- ID ----------
        id_pattern = re.compile(r"ID:([0-9A-Fa-f]{4})")
        for row_idx, ln in enumerate(plain_lines):
            m = id_pattern.search(ln)
            if m:
                values[0] = m.group(1).upper()
                term_hex = self._row_hex_color(color_matrix, row_idx)
                colors[0] = _excel_color_from_hex(term_hex)
                break

        # ---------- Остальные поля ----------
        # (label, column_index, mode)
        rules = [
            ("UART", 1, "bracket"),
            ("Voltage", 2, "bracket"),
            ("U_bat", 3, "number"),
            ("U_src", 4, "number"),
            ("Current", 5, "bracket"),
            ("I_crg", 6, "number"),
            ("I_ch1", 7, "number"),
            ("I_ch2", 8, "number"),
            ("Charger", 9, "bracket"),
            ("M_sens", 10, "bracket"),
            ("L_sens", 11, "bracket"),
        ]

        for row_idx, ln in enumerate(plain_lines):
            if not ln.strip():
                continue

            for label, col_idx, mode in rules:
                if label not in ln:
                    continue

                term_hex = self._row_hex_color(color_matrix, row_idx)
                colors[col_idx] = _excel_color_from_hex(term_hex)

                if mode == "bracket":
                    m = re.search(r"\[([^\]]*)\]", ln)
                    if m:
                        values[col_idx] = m.group(1).strip()

                elif mode == "number":
                    m = re.search(rf"{re.escape(label)}\s+(-?\d+)", ln)
                    if m:
                        values[col_idx] = m.group(1).strip()

        return values, colors

    # ----------------------------------------------------------
    # Сохранение блока
    # ----------------------------------------------------------
    def save_block(
        self,
        lines: List[str],
        color_matrix=None,
        short_id: Optional[str] = None,
        auto: bool = False,
    ) -> None:
        """
        Сохраняем блок:
        - lines        — строки pyte (с ANSI, мы сами очистим)
        - color_matrix — матрица цветов CanvasTerminal.last_colors
        - short_id     — ID устройства (CRC16 UID), если есть
        - auto         — True, если автосохранение по PASSED

        Логика:

        * PASSED ищется по ВСЕМ строкам кадра.
        * В auto-режиме:
            - если PASSED нет — вообще ничего не пишем;
            - если PASSED есть, но ID пустой — ничего не пишем;
            - если PASSED есть и этот ID уже был — ничего не пишем;
            - если PASSED есть и ID новый — пишем TXT + лист PASSED.
        * В ручном режиме (auto=False):
            - если в кадре есть PASSED — пишем TXT + PASSED;
            - иначе — TXT + основной лист Sheet.
        """
        if not lines:
            self._set_status("MPPT: нет блока для сохранения", "red")
            return

        ts = timestamp_str()
        plain = [strip_ansi(l) for l in lines]

        # ---------- Проверка PASSED по ВСЕМ строкам ----------
        is_passed = any("PASSED" in l.upper() for l in plain)

        # парсим кадр
        values, colors = self._parse_frame(lines, color_matrix)

        # short_id приоритетнее ID из кадра
        if short_id:
            values[0] = short_id.upper()
            colors[0] = "000000"

        cur_id = values[0].strip()

        # ---------- Решение: сохранять или нет, и в какой лист ----------
        if auto:
            # авто-режим: сохраняем только если PASSED и есть нормальный ID, и это не дубль
            if not is_passed:
                return

            if not cur_id:
                self._set_status(
                    "MPPT: PASSED без ID — авто-сохранение пропущено", "yellow"
                )
                return

            if self.last_passed_id == cur_id:
                # дубль — тихо пропускаем
                return

            self.last_passed_id = cur_id
            target_sheet_name = "PASSED"
        else:
            # ручной режим
            target_sheet_name = "PASSED" if is_passed else "Sheet"

        # ---------- TXT LOG (пишем только если решили сохранять) ----------
        os.makedirs(os.path.dirname(self.txt_path), exist_ok=True)
        with open(self.txt_path, "a", encoding="utf-8") as f:
            f.write("\n" + "-" * 50 + "\n")
            f.write(f"[{ts}]\n")
            for l in plain:
                f.write(l.rstrip() + "\n")
            f.write("-" * 50 + "\n")

        # ---------- Excel ----------
        wb, ws, ws_passed = self._ensure_workbook()
        target_ws = ws_passed if target_sheet_name == "PASSED" else ws

        row_idx = target_ws.max_row + 1
        for i, (val, col) in enumerate(zip(values, colors), start=1):
            cell = target_ws.cell(row=row_idx, column=i, value=val)
            cell.font = Font(color=col)

        try:
            wb.save(self.xlsx_path)
        except PermissionError:
            self._set_status(
                f"MPPT: не удалось сохранить Excel (файл открыт?): {self.xlsx_path}",
                "red",
            )
            return

        # финальные статусы
        if target_sheet_name == "PASSED":
            self._set_status(
                f"MPPT: сохранено в лист PASSED (строка {row_idx})", "green"
            )
        else:
            self._set_status(f"MPPT: блок сохранён (строка {row_idx})", "green")

    # ----------------------------------------------------------
    # Git-поддержка
    # ----------------------------------------------------------
    def _is_git_repo(self) -> bool:
        """Проверяет, есть ли .git у логов."""
        return os.path.isdir(os.path.join(self.logs_dir, ".git"))


    def _run_git(self, *args) -> subprocess.CompletedProcess:
        """Запуск git-команды в logs_dir с захватом вывода."""
        proc = subprocess.run(
            args,
            cwd=self.logs_dir,
            capture_output=True,
            text=True,
        )
        return proc


    def _ensure_git_repo(self) -> None:
        """
        Конфигурирует репозиторий:
        - git init (если нет .git)
        - git remote add origin (если нет origin)
        - git fetch (+retry)
        - git branch --set-upstream-to=origin/main/master
        - git pull --ff-only (+retry)
        """

        ORIGIN_URL = "http://dis-electronics:30000/scheck/swmpptv7_10a_rev1_reject.git"

        # ----------------------------
        # 1) Если .git нет → git init
        # ----------------------------
        if not self._is_git_repo():
            self._set_git_status("Git: init нового репозитория…", "#85c1ff")

            proc = self._run_git("git", "init")
            if proc.returncode != 0:
                self._set_git_status(f"Git: init ошибка: {proc.stderr}", "red")
                return

        # ----------------------------
        # 2) Проверка наличия origin
        # ----------------------------
        proc = self._run_git("git", "remote")
        remotes = proc.stdout.split()
        has_origin = "origin" in remotes

        if not has_origin:
            self._set_git_status("Git: добавление origin…", "#85c1ff")
            proc = self._run_git("git", "remote", "add", "origin", ORIGIN_URL)
            if proc.returncode != 0:
                self._set_git_status(f"Git: ошибка добавления origin: {proc.stderr}", "red")
                return

        # ----------------------------
        # 3) FETCH (c retry)
        # ----------------------------
        for attempt in range(1, 4):
            proc = self._run_git("git", "fetch")
            if proc.returncode == 0:
                break

            if attempt == 3:
                self._set_git_status(
                    f"Git: fetch не удался после 3 попыток: {proc.stderr}", "red"
                )
                return

            self._set_git_status(
                f"Git: fetch ошибка, повтор {attempt}/3: {proc.stderr}", "yellow"
            )
            time.sleep(2)

        # ----------------------------
        # 4) Установка upstream
        # ----------------------------
        upstream_branch = None

        proc = self._run_git("git", "branch", "--set-upstream-to=origin/main")
        if proc.returncode == 0:
            upstream_branch = "main"
            self._set_git_status("Git: upstream → origin/main", "green")
        else:
            proc = self._run_git("git", "branch", "--set-upstream-to=origin/master")
            if proc.returncode == 0:
                upstream_branch = "master"
                self._set_git_status("Git: upstream → origin/master", "green")
            else:
                self._set_git_status(
                    "Git: не удалось установить upstream (ни main, ни master)",
                    "yellow",
                )

        # Без upstream pull невозможен
        if not upstream_branch:
            return

        # ----------------------------
        # 5) ПЕРВЫЙ PULL (c retry)
        # ----------------------------
        for attempt in range(1, 4):
            proc = self._run_git("git", "pull", "--ff-only")

            if proc.returncode == 0:
                self._set_git_status("Git: первый pull завершён", "green")
                return

            if "fast-forward" in proc.stderr.lower():
                self._set_git_status(
                    "Git: FF pull невозможен (разные коммиты)", "yellow"
                )
                return

            if attempt == 3:
                self._set_git_status(
                    f"Git: первый pull ошибка: {proc.stderr}", "red"
                )
                return

            self._set_git_status(
                f"Git: pull ошибка, повтор {attempt}/3: {proc.stderr}",
                "yellow",
            )
            time.sleep(2)

    def _git_pull_on_start(self, retries: int = 3, delay: float = 2.0) -> None:
        """Fetch + pull --ff-only с 3 попытками и корректными статусами."""

        # ----------------------------
        # 1) F E T C H  (с retry)
        # ----------------------------
        for attempt in range(1, retries + 1):
            proc = self._run_git("git", "fetch")

            if proc.returncode == 0:
                break

            if attempt == retries:
                self._set_git_status(
                    f"Git: fetch не удался: {proc.stderr}", "red"
                )
                return

            self._set_git_status(
                f"Git: fetch ошибка ({attempt}/{retries}), повтор через {delay}с",
                "yellow",
            )
            time.sleep(delay)

        # ----------------------------
        # 2) P U L L  (с retry)
        # ----------------------------
        for attempt in range(1, retries + 1):
            proc = self._run_git("git", "pull", "--ff-only")

            stdout = proc.stdout.lower()
            stderr = proc.stderr.lower()

            # УСПЕХ
            if proc.returncode == 0:
                if "up to date" in stdout:
                    self._set_git_status("Git: уже актуально (up to date)", "green")
                else:
                    self._set_git_status("Git: pull завершён", "green")
                return

            # FF невозможен
            if "fast-forward" in stderr:
                self._set_git_status(
                    "Git: FF pull невозможен (разные коммиты)", "yellow"
                )
                return

            # ошибка (retry)
            if attempt == retries:
                self._set_git_status(
                    f"Git: pull ошибка: {proc.stderr}", "red"
                )
                return

            self._set_git_status(
                f"Git: pull ошибка ({attempt}/{retries}), повтор…",
                "yellow",
            )
            time.sleep(delay)



    def _git_pull_on_start_ui(self) -> None:
        """UI-обёртка вокруг git pull."""
        try:
            self._git_pull_on_start()
        except Exception as e:
            self._set_git_status(f"Git: ошибка pull: {e}", "red")


    def git_commit_logs(self) -> None:
        """
        Добавить изменения в git и сделать commit.
        """
        if not self._is_git_repo():
            self._set_git_status(
                "Git: каталог логов не является репозиторием", "yellow"
            )
            return

        try:
            self._set_git_status("Git: commit…", "#85c1ff")
            self._run_git("git", "add", ".")
            msg = f"Auto log commit {timestamp_str()}"
            self._run_git("git", "commit", "-m", msg)
            self._set_git_status("Git: commit завершён", "green")
        except RuntimeError as e:
            text = str(e)
            if "nothing to commit" in text:
                self._set_git_status("Git: нечего коммитить", "yellow")
            else:
                self._set_git_status(f"Git: ошибка commit: {text}", "red")
        except Exception as e:
            self._set_git_status(f"Git: общая ошибка commit: {e}", "red")

    def git_push(self) -> None:
        """
        Выполнить git push.
        """
        if not self._is_git_repo():
            self._set_git_status(
                "Git: каталог логов не является репозиторием", "yellow"
            )
            return

        try:
            self._set_git_status("Git: push…", "#85c1ff")
            self._run_git("git", "push")
            self._set_git_status("Git: push завершён", "green")
        except Exception as e:
            self._set_git_status(f"Git: ошибка push: {e}", "red")
